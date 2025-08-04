import cv2
import face_recognition
import pickle
import os
import numpy as np
from datetime import datetime
from openpyxl import Workbook, load_workbook

ENCODING_FILE = "../encodings/face_encodings.pkl"
ATTENDANCE_FILE = "../attendance/attendance.csv"
EXCEL_FILE = "../attendance/attendance.xlsx"

def mark_attendance(name):
    now = datetime.now()
    dt_string = now.strftime('%Y-%m-%d %H:%M:%S')

    # Ensure folder exists
    os.makedirs(os.path.dirname(ATTENDANCE_FILE), exist_ok=True)

    # Step 1: Write to CSV
    if not os.path.exists(ATTENDANCE_FILE):
        with open(ATTENDANCE_FILE, 'w') as f:
            f.write("Name,Time\n")

    with open(ATTENDANCE_FILE, 'r+') as f:
        existing_entries = f.readlines()
        names_logged = [line.split(',')[0] for line in existing_entries[1:]]
        if name not in names_logged:
            f.write(f"{name},{dt_string}\n")
            print(f"[INFO] Marked {name} present at {dt_string}")

    # Step 2: Write to Excel
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        if "Attendance" in wb.sheetnames:
            ws = wb["Attendance"]
        else:
            ws = wb.create_sheet("Attendance")
            ws.append(["Name", "Time"])
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(["Name", "Time"])

    # Avoid duplicate entries
    names_in_excel = [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]
    if name not in names_in_excel:
        ws.append([name, dt_string])
        wb.save(EXCEL_FILE)

def recognize_faces():
    print("[INFO] Loading encodings...")
    with open(ENCODING_FILE, 'rb') as f:
        data = pickle.load(f)

    print("[INFO] Starting webcam...")
    cap = cv2.VideoCapture(0)

    if not cap.isOpened():
        print("[ERROR] Could not open webcam")
        return

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

        boxes = face_recognition.face_locations(rgb_small_frame)
        encodings = face_recognition.face_encodings(rgb_small_frame, boxes)

        names = []

        for encoding in encodings:
            matches = face_recognition.compare_faces(data["encodings"], encoding)
            name = "Unknown"

            face_distances = face_recognition.face_distance(data["encodings"], encoding)
            best_match_index = np.argmin(face_distances)

            if matches[best_match_index]:
                name = data["names"][best_match_index]
                mark_attendance(name)

            names.append(name)

        for (top, right, bottom, left), name in zip(boxes, names):
            top *= 4
            right *= 4
            bottom *= 4
            left *= 4

            cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
            cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 255, 0), cv2.FILLED)
            font = cv2.FONT_HERSHEY_SIMPLEX
            cv2.putText(frame, name, (left + 6, bottom - 6), font, 1, (0, 0, 0), 2)

        cv2.imshow('Face Recognition Attendance', frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

if __name__ == "__main__":
    recognize_faces()
