# app.py

from flask import Flask, render_template, Response
import cv2
import face_recognition
import pickle
import os
import numpy as np
from datetime import datetime
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

# Paths
ENCODING_FILE = "encodings/face_encodings.pkl"
CSV_FILE = "attendance/attendance.csv"
XLSX_FILE = "attendance/attendance.xlsx"

# Load encodings once
with open(ENCODING_FILE, 'rb') as f:
    data = pickle.load(f)

# Attendance logging function
def mark_attendance(name):
    now = datetime.now()
    dt_string = now.strftime('%Y-%m-%d %H:%M:%S')

    os.makedirs(os.path.dirname(CSV_FILE), exist_ok=True)

    # CSV Logging
    if not os.path.exists(CSV_FILE):
        with open(CSV_FILE, 'w') as f:
            f.write("Name,Time\n")

    with open(CSV_FILE, 'r+') as f:
        lines = f.readlines()
        names_logged = [line.split(',')[0] for line in lines[1:]]
        if name not in names_logged:
            f.write(f"{name},{dt_string}\n")

    # Excel Logging
    if os.path.exists(XLSX_FILE):
        wb = load_workbook(XLSX_FILE)
        ws = wb["Attendance"] if "Attendance" in wb.sheetnames else wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(["Name", "Time"])

    names_in_excel = [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]
    if name not in names_in_excel:
        ws.append([name, dt_string])
        wb.save(XLSX_FILE)

# Generator function for video feed
def gen_frames():
    cap = cv2.VideoCapture(0)

    if not cap.isOpened():
        raise RuntimeError("Could not open webcam")

    while True:
        success, frame = cap.read()
        if not success:
            break

        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

        boxes = face_recognition.face_locations(rgb_small_frame)
        encodings = face_recognition.face_encodings(rgb_small_frame, boxes)

        names = []
        for encoding in encodings:
            matches = face_recognition.compare_faces(data["encodings"], encoding)
            name = "Unknown"

            face_distances = face_recognition.face_distance(data["encodings"], encoding)
            best_match_index = np.argmin(face_distances)

            if matches[best_match_index]:
                name = data["names"][best_match_index]
                mark_attendance(name)

            names.append(name)

        for (top, right, bottom, left), name in zip(boxes, names):
            top *= 4
            right *= 4
            bottom *= 4
            left *= 4

            cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
            cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 255, 0), cv2.FILLED)
            cv2.putText(frame, name, (left + 6, bottom - 6), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 0), 2)

        ret, buffer = cv2.imencode('.jpg', frame)
        frame = buffer.tobytes()

        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/video')
def video():
    return Response(gen_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

if __name__ == '__main__':
    app.run(debug=True)
